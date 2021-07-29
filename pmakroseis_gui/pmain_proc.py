"""
Макросейсмика, 2020,
главный модуль
"""
# pmain_proc

import numpy as np
from geogr_distance import *
from pinp_proc import *

import pfunct
# from pinp_struct import *
import pinp_struct
import ptest_alg

from pnumpy import *
import math
import numba
from scipy.optimize import minimize
import copy
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font, RichTextProperties, RegularTextRun
from openpyxl.styles import Font, Fill
# from psort import *
# from pfit import *
from tkinter import messagebox as mb
from ptkinter_menu_proc import *

log_file_name: str
gran = 0.001  # 0.001% изменения функции по абсолютной величине
# для analyze_diff, порог для max_var/max_fun, возможные уровни срабатывания 1, 10, 20, 100, 400


def main_proc(is_view1: bool) -> None:
    global log_file_name
    if test_mode:
        if is_view1:
            print('тестовый режим')                # txt_input работает, когда inf_input = False
        (file_name, log_file_name) = prepare_input_filename(inf_input=True, txt_input=True, is_view=is_view1)
        if is_view1:
            print(file_name, log_file_name)
        good_end = the_input(file_name, is_view1)
        if not good_end:
            print('Ошибка в данных, программа расчеты не выполняла')
        work_with_data(is_view1)

    print('Нормальное завершение')


def work_with_data(is_view1: bool) -> (bool, int, float, float, float, float, float, list):
    """
    Подготовка к работе и минимизация
    bool
    """
    row: int; i: int;

    if is_view1: print('work_with_data')
    (the_dict, the_arr) = pinp_struct.get_dat_struct(pinp_struct.curr_nstruct)
    row = the_dict["npoint"]
    lat_arr = the_arr[:, 0]  # Lat
    # print('lat_arr ', np.size(lat_arr)); print(lat_arr)
    lon_arr = the_arr[:, 1]  # Lat
    # print('lon_arr ', np.size(lon_arr)); print(lon_arr)
    h_arr = the_arr[:, 2]/1000  # Alt переводим в км
    # print('h_arr ', np.size(h_arr)); print(h_arr)
    i_fact_arr = the_arr[:, 3]
    npoint = np.size(i_fact_arr)
    eqval = (row == npoint)
    di_arr = the_arr[:, 4]
    # print('i_fact_arr ', np.size(i_fact_arr)); print(i_fact_arr)
    if not eqval:
        # заглушка для данных
        num, lat_, lon_, dep_, mag_, fun_ = -13, -13, -13, -13, -13, -13,
    else:
        (num, lat_, lon_, dep_, mag_, fun_, res_list_) = minimize_func(npoint, lat_arr, lon_arr, h_arr, i_fact_arr, di_arr)
    return eqval, num, lat_, lon_, dep_, mag_, fun_, res_list_
    # view_2d_array(dat, nrow: int, ncol: int, test_info='Просмотр в work_with_data '):


def calc_diff(iii: int, x0: float, x1: float, x2: float, x3: float, f: float, result_list: list) -> \
                                                        (float, float, float, float, float, float):
                                                        # dx0    dx1   dx2    dx3,   df,    gran
    if iii == 0:
        return math.nan, math.nan, math.nan, math.nan, math.nan, math.nan
    else:
        n = 100
        dat = result_list[iii-1]
        dx0 = ((x0 - dat[0])/dat[0])*n  # lat_ = x0[0]
        dx1 = ((x1 - dat[1])/dat[1])*n  # lon_ = x0[1]
        dx2 = ((x2 - dat[2])/dat[2])*n  # dep_ = x0[2]
        dx3 = ((x3 - dat[3])/dat[3])*n  # mag_ = x0[3]
        df  = ((f  - dat[4])/dat[4])*n  # f = x0[4]
        dgran_ = get_diffgran(dx0, dx1, dx2, dx3, df)
        return (dx0, dx1, dx2, dx3, df, dgran_)


def get_diffgran(dx0: float, dx1: float, dx2: float, dx3: float, df: float) -> float:
    # max_dvar = abs(max(dx0, dx1, dx2, dx3))
    max_dvar = max(abs(dx0), abs(dx1), abs(dx2), abs(dx3))
    max_dfun = abs(df)
    # max_dvar/max_dfun - числа не более 1, 10, 40 - как бы "линейная" зависимость
    # max_dfun/max_dvar - как бы парабола, ветви вверх, аппроксимируем, находим индекс минимума
    # как бы 2-3 минимума, но они после больших значений, так что реагировать будут на первый
    return max_dvar/max_dfun


def analyze_diff(iii: int, result_list: list) -> bool:
    """
    Определяет как быстро меняется самая вариабельная переменная относительно изменения функции
    Если более gran, то return True / иначе return False
    """
    global gran
    if iii != 0:   # 5    6     7   8     9
        dat = result_list[iii]
        # max_var = max(abs(dat[5]), abs(dat[6]), abs(dat[7]), abs(dat[8]))
        max_fun = abs(dat[9])
        # d = max_var/max_fun  # как быстро меняется самая вариабельная переменная относительно изменения функции
        d = max_fun  # как быстро меняется целевая функция
        # print('max_var/max_fun = %7.5f' %  d)
        if d > gran:  # если  - числа не более 1, 10, 40 - как бы "линейная" зависимость
            # print('d > gran')
            return True
        else:
            return False


def output_txt_res(fname: str, result_list: list) -> None:
    """
    Неполный вариант по выводим данным, полный см. output_xls_res
    Вывод в txt файл итераций минимизации:
    номер, lat, lon, dep, mag, fun
    """
    fff = open(fname,  mode='w')  # открытие log-файла
    llen = len(result_list)
    fff.write('{0:6s} {1:9s} {2:9s} {3:9s} {4:9s} {5:11s} {6:9s} {7:9s} {8:9s} {9:9s} {10:11s} {11:11s}\n'.format('N', 'Lat', 'Lon', 'Dep', 'Mag', 'Fun', 'dLat,%', 'dLon,%', 'dDep,%', 'dMag,%', 'dFun,%', 'max_var,%/dFun,%'))
    for i in range(llen):
        d = result_list[i]
        fff.write('{0:5d} {1:9.5f} {2:9.5f} {3:9.5f} {4:9.5f} {5:11.5f} {6:9.5f} {7:9.5f} {8:9.5f} {9:9.5f}  {10:11.5f} {11:11.5f}\n'.format(i, d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7], d[8], d[9], d[10]))
    fff.close()


def calc_imod(a_: float, b_: float, c_: float, lat: float, lon: float, dep: float, mag: float) -> np.ndarray:
    (the_dict, the_arr) = pinp_struct.get_dat_struct(pinp_struct.curr_nstruct)
    row = the_dict["npoint"]
    # print('row =', row)
    imod = np.random.random(row)
    lat_arr = the_arr[:, 0]  # Lat
    # print('lat_arr ', np.size(lat_arr)); print(lat_arr)
    lon_arr = the_arr[:, 1]  # Lat
    # print('lon_arr ', np.size(lon_arr)); print(lon_arr)
    h_arr = the_arr[:, 2]/1000  # Alt переводим в км
    # print('h_arr ', np.size(h_arr)); print(h_arr)
    i_fact_arr = the_arr[:, 3]
    npoint = np.size(i_fact_arr)
    eqval = (row == npoint)
    for i in range(row):
        (epi_len, hypo_len) = pinp_struct.calc_distance(lat_arr[i], lon_arr[i], h_arr[i], lat, lon, dep)
        dist3 = hypo_len
        imod[i] = pinp_struct.makroseis_fun(a=a_, b=b_, c=c_, dist=dist3, mag=mag,
                                            type_of_macro_fun_=pinp_struct.type_of_macro_fun)
    return imod

def calc_imod2(a_: float, b_: float, c_: float, lat_pnt: float, lon_pnt: float, dep: float, mag: float,
                          lat_arr: np.ndarray, lon_arr: np.ndarray, h_arr1: np.ndarray) -> (np.ndarray, np.ndarray):
    row = len(lat_arr)
    imod_arr = np.random.random(row)
    hypo_len_arr = np.random.random(row)
    epi_len_arr = np.random.random(row)
    h_arr = h_arr1/1000  # Alt переводим в км
    for i in range(row):
        (epi_len, hypo_len) = pinp_struct.calc_distance(lat_arr[i], lon_arr[i], h_arr[i], lat_pnt, lon_pnt, dep)
        epi_len_arr[i] = epi_len
        hypo_len_arr[i] = hypo_len
        dist3 = hypo_len
        imod_arr[i] = pinp_struct.makroseis_fun(a=a_, b=b_, c=c_, dist=dist3, mag=mag,
                                            type_of_macro_fun_=pinp_struct.type_of_macro_fun)
    return imod_arr, hypo_len_arr, epi_len_arr


def prepare_output_xlsres_fn(resdir:str, inffname:str) -> str:
    fn = gfn(inffname) + '_res.xlsx'
    s = "\\".join([resdir, fn])
    return s

def prepare_output_xlspointlist_fn(resdir:str, inffname:str) -> str:
    fn = gfn(inffname) + '_pointlist.xlsx'
    s = "\\".join([resdir, fn])
    return s


def output_xlspointlist_stage2(fname: str, arr_dep: np.ndarray, arr_mag: np.ndarray, res_matr_lin: np.ndarray,
                               correct_fun: float, info_str:str) -> bool:
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    my_sheet.title = "Глуб., магн. и ц.функ., 2 этап"
    llen = len(arr_dep)

    #---- Сортировка
    ind = np.argsort(res_matr_lin)
    sort_dep = arr_dep[ind]
    sort_mag = arr_mag[ind]
    sort_fun = res_matr_lin[ind]
    D1 = llen // 10   # число точек в 1 дециле
    my_sheet.cell(row=2, column= 1).value = 'Номер'
    my_sheet.cell(row=2, column= 2).value = 'Глубина, км'
    my_sheet.cell(row=2, column= 3).value = 'Магнитуда'
    my_sheet.cell(row=2, column= 4).value = 'Целевая функция'
    rrr = range(1, 5)
    for i in range(llen):
        rrow = i + 3
        my_sheet.cell(row=rrow, column=1).value = i+1
        my_sheet.cell(row=rrow, column=2).value = sort_dep[i]
        my_sheet.cell(row=rrow, column=3).value = sort_mag[i]
        my_sheet.cell(row=rrow, column=4).value = round(sort_fun[i], 6)
        # https://ru.stackoverflow.com/questions/1189595/Как-поменять-цвет-в-ячейке-с-помощью-openpyxl
        # https://lavrynenko.com/openpyxl-kak-menyat-shrift-i-cvet-v-yachejke/
        if (abs(correct_fun - sort_fun[i]) < 1e-10):
            for j in rrr:
                my_sheet.cell(row=rrow, column=j).font = openpyxl.styles.Font(color='FF0000', bold=True)
            i_sav = i
            my_sheet.cell(row=rrow, column=5).value = info_str
            my_sheet.cell(row=rrow, column=5).font = openpyxl.styles.Font(color='FF0000', bold=True)
        else:
            if i < D1: # форматируем для точек в 1 дециле
                for j in rrr:
                    my_sheet.cell(row=rrow, column=j).font = openpyxl.styles.Font(color='312BED', bold=True)
    #----- Первая строка---- начало
    my_sheet.cell(row=1, column=1).value = i_sav + 1
    my_sheet.cell(row=1, column=2).value = sort_dep[i_sav]
    my_sheet.cell(row=1, column=3).value = sort_mag[i_sav]
    my_sheet.cell(row=1, column=4).value = round(sort_fun[i_sav], 6)
    for j in rrr:
        my_sheet.cell(row=1, column=j).font = openpyxl.styles.Font(color='FF0000', bold=True)
    my_sheet.cell(row=1, column=5).value = info_str
    my_sheet.cell(row=1, column=5).font = openpyxl.styles.Font(color='FF0000', bold=True)
    #----- Первая строка---- конец

    my_sheet.column_dimensions['B'].width = 12
    my_sheet.column_dimensions['C'].width = 12
    my_sheet.column_dimensions['D'].width = 19
    try:
        my_wb.save(fname)
        bres = True
    except Exception:
        bres = False
    return bres


def prepare_and_output_xlsres_stage2(resdir:str, inffname:str, sq_name:str,
                                     lat: np.ndarray, lon: np.ndarray, alt: np.ndarray, ifact: np.ndarray,
                                     di: np.ndarray, N: np.ndarray, the_name: np.ndarray, dist: np.ndarray,
                                     imod_min: np.ndarray, imod_corr: np.ndarray,
                                     ini_dat: list, res_dat: list,
                                     sort_dist_min: np.ndarray, sort_imod_min: np.ndarray,
                                     sort_dist_corr: np.ndarray, sort_imod_corr: np.ndarray,
                                     sort_ifact_arr: np.ndarray, sort_name: np.ndarray,
                                     min_icorrect:
                                     int, arr_dep: np.ndarray, arr_mag: np.ndarray, res_matr_lin: np.ndarray,
                                     info_str:str) -> (bool, bool):
    fname_res = prepare_output_xlsres_fn(resdir, inffname)
    #----------------- xls_res
    out1_res = output_xls_res_stage2(fname_res, sq_name, lat, lon, alt , ifact, di, N, the_name, dist,
                      imod_min, imod_corr, ini_dat, res_dat,
                                     sort_dist_min, sort_imod_min,
                                     sort_dist_corr, sort_imod_corr, sort_ifact_arr, sort_name)

    #----------------- xls_pointlist
    # def output_xlspointlist_stage2(fname: str, arr_dep: np.ndarray, arr_mag: np.ndarray, res_matr_lin: np.ndarray) -> bool:
    fname_pl = prepare_output_xlspointlist_fn(resdir, inffname)
    correct_fun = res_matr_lin[min_icorrect]
    out2_res = output_xlspointlist_stage2(fname_pl, arr_dep, arr_mag, res_matr_lin, correct_fun, info_str)
    return (True, out2_res)


def output_xls_res_stage2(fname:str, sq_name:str, lat: np.ndarray, lon: np.ndarray, alt: np.ndarray, ifact: np.ndarray, di: np.ndarray, N: np.ndarray,
                          the_name: np.ndarray, dist: np.ndarray, imod_min: np.ndarray, imod_corr: np.ndarray,
                          ini_dat: list, res_dat: list,
                          sort_dist_min: np.ndarray, sort_imod_min: np.ndarray,
                          sort_dist_corr: np.ndarray, sort_imod_corr: np.ndarray,
                          sort_ifact_arr: np.ndarray, sort_name: np.ndarray) -> bool:
    """
    Вывод в xlsx файл исходных данных и результаитов минимизации:
    lat, lon, alt, ifact, di, N, the_name, imod
    """
    ################### Первый лист
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    my_sheet.title = "Данные и Imod"
    llen = len(lat)
    my_sheet.cell(row=1, column= 1).value = 'Lat'
    my_sheet.cell(row=1, column= 2).value = 'Lon'
    my_sheet.cell(row=1, column= 3).value = 'Alt'
    my_sheet.cell(row=1, column= 4).value = 'Ifact'
    my_sheet.cell(row=1, column= 5).value = 'dI'
    my_sheet.cell(row=1, column= 6).value = 'N'
    my_sheet.cell(row=1, column= 7).value = 'Нас.пункт'
    my_sheet.cell(row=1, column= 8).value = 'Imod_corr'
    my_sheet.cell(row=1, column= 9).value = 'В пределах dI'
    my_sheet.cell(row=1, column=10).value = 'От эпицентра (выбр.точка), км'
    my_sheet.cell(row=1, column=11).value = 'Imod (Fmin)'
    dist_on_surf = ptest_alg.calc_dist_onsurf_full(res_dat[1], res_dat[2], lat, lon)
    for i in range(llen):
        rrow = i + 2
        my_sheet.cell(row=rrow, column=1).value = lat[i]
        my_sheet.cell(row=rrow, column=2).value = lon[i]
        my_sheet.cell(row=rrow, column=3).value = alt[i]
        my_sheet.cell(row=rrow, column=4).value = ifact[i]
        my_sheet.cell(row=rrow, column=5).value = di[i]
        my_sheet.cell(row=rrow, column=6).value = N[i]
        my_sheet.cell(row=rrow, column=7).value = the_name[i]
        my_sheet.cell(row=rrow, column=8).value = round(imod_corr[i], 4)
        my_sheet.cell(row=rrow, column=9).value = pfunct.get_xls_bool(abs(ifact[i]-imod_corr[i]) <= (di[i]/2))
        my_sheet.cell(row=rrow, column=10).value = round(dist[i], 4)
        my_sheet.cell(row=rrow, column=11).value = round(imod_min[i], 4)

    my_sheet.column_dimensions['G'].width = 25
    my_sheet.column_dimensions['H'].width = 15
    my_sheet.column_dimensions['I'].width = 15
    my_sheet.column_dimensions['J'].width = 35
    my_sheet.column_dimensions['K'].width = 15

    ################### Второй лист
    # https://ru.stackoverflow.com/questions/828957/С-помощью-python-cоздать-новый-лист-в-уже-имеющемся-excel-файле
    res_sheet = my_wb.create_sheet('Результаты')
    res_sheet.cell(row=1, column=1).value = sq_name
    res_sheet.cell(row=2, column=1).value = 'Широта';     res_sheet.cell(row=2, column=2).value = ini_dat[0]
    res_sheet.cell(row=3, column=1).value = 'Долгота';    res_sheet.cell(row=3, column=2).value = ini_dat[1]

    res_sheet.cell(row=5, column=1).value = 'Начальное приближение'
    res_sheet.cell(row=6, column=1).value = 'Глубина';    res_sheet.cell(row=6, column=2).value = ini_dat[2]
    res_sheet.cell(row=7, column=1).value = 'Магнитуда';  res_sheet.cell(row=7, column=2).value = ini_dat[3]

    res_sheet.cell(row=9, column=1).value = 'Наименьшее значение целевой функции'
    res_sheet.cell(row=10, column=1).value = 'Глубина';    res_sheet.cell(row=10, column=2).value = res_dat[0]
    res_sheet.cell(row=11, column=1).value = 'Магнитуда';  res_sheet.cell(row=11, column=2).value = res_dat[1]
    res_sheet.cell(row=12, column=1).value = 'Значение целевой функции';
    res_sheet.cell(row=12, column=2).value =  res_dat[2]

    res_sheet.cell(row=14, column=1).value = 'Выбранный результат минимизации'
    res_sheet.cell(row=15, column=1).value = 'Глубина';   res_sheet.cell(row=15, column=2).value = res_dat[3]
    res_sheet.cell(row=16, column=1).value = 'Магнитуда'; res_sheet.cell(row=16, column=2).value = res_dat[4]
    res_sheet.cell(row=17, column=1).value = 'Значение целевой функции'
    res_sheet.cell(row=17, column=2).value = res_dat[5]
    res_sheet.cell(row=18, column=2).value = res_dat[6]
                   # http://espressocode.top/python-adjusting-rows-and-columns-of-an-excel-file-using-openpyxl-module/
    res_sheet.column_dimensions['A'].width = 40
    # ################### Третий лист
    chart_sheet = my_wb.create_sheet('Диаграмма')
    chart_sheet.cell(row=1, column=1).value = 'От гипоцентра выбр, км'
    chart_sheet.cell(row=1, column=2).value = 'I_fact'
    chart_sheet.cell(row=1, column=3).value = 'I_mod_выбр'
    chart_sheet.cell(row=1, column=4).value = 'Нас.пункт'
    chart_sheet.cell(row=1, column=5).value = 'От гипоцентра fmin, км'
    chart_sheet.cell(row=1, column=6).value = 'I_mod_fmin'
    nnn = len(imod_min)
    for i in range(nnn):
        rrow = i + 2
        chart_sheet.cell(row=rrow, column=1).value = sort_dist_corr[i]
        chart_sheet.cell(row=rrow, column=2).value = sort_ifact_arr[i]
        chart_sheet.cell(row=rrow, column=3).value = sort_imod_corr[i]
        chart_sheet.cell(row=rrow, column=4).value = sort_name[i]
        chart_sheet.cell(row=rrow, column=5).value = sort_dist_min[i]
        chart_sheet.cell(row=rrow, column=6).value = sort_imod_min[i]

    chart_sheet.column_dimensions['A'].width = 24
    chart_sheet.column_dimensions['B'].width = 10
    chart_sheet.column_dimensions['C'].width = 15
    chart_sheet.column_dimensions['D'].width = 22
    chart_sheet.column_dimensions['E'].width = 16
    chart_sheet.column_dimensions['F'].width = 16

    chart = ScatterChart()
    # создать данные для построения графика
    x1values = Reference(chart_sheet, min_col=1, min_row=2, max_row=nnn+2)
    y10values = Reference(chart_sheet, min_col=2, min_row=2, max_row=nnn+2)
    y11values = Reference(chart_sheet, min_col=3, min_row=2, max_row=nnn+2)

    x2values = Reference(chart_sheet, min_col=5, min_row=2, max_row=nnn+2)
    y2values = Reference(chart_sheet, min_col=6, min_row=2, max_row=nnn+2)

    # создать серии данных
    series = Series(values=y10values, xvalues=x1values,  title="Исходная") # zvalues=size,
    series.graphicalProperties.line.solidFill = '0000FF'
    series.graphicalProperties.line.width = 25000
    series1 = Series(values=y11values, xvalues=x1values, title="Рассчитанная, выбр.") # zvalues=size
    series1.graphicalProperties.line.solidFill = 'FF0000'
    series1.graphicalProperties.line.width = 25000
    series2 = Series(values=y2values, xvalues=x2values, title="Рассчитанная, fmin.") # zvalues=size
    series2.graphicalProperties.line.solidFill = '006400'
    series2.graphicalProperties.line.width = 25000

    font_test = openpyxl.drawing.text.Font(typeface='Arial')
    cp = CharacterProperties(latin=font_test, sz=1500)
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    min_x = min(np.min(sort_dist_min), np.min(sort_dist_corr))
    max_x = max(np.max(sort_dist_min), np.max(sort_dist_corr))
    chart.x_axis.scaling.min = round(math.floor(min_x),-1) - 10
    chart.x_axis.scaling.max = round(math.ceil(max_x) ,-1) + 10
    # chart.y_axis.title.

    min_y = min(np.min(sort_ifact_arr), np.min(sort_imod_min), np.min(sort_imod_corr))
    max_y = max(np.max(sort_ifact_arr), np.max(sort_imod_min), np.max(sort_imod_corr))
    chart.y_axis.scaling.min = round(math.floor(min_y), 0) - 1
    chart.y_axis.scaling.max = round(math.ceil(max_y),  0) + 1
    chart.legend.position = 'b'
    #
    # добавить данные серии в объект диаграммы
    chart.series.append(series)
    chart.series.append(series1)
    chart.series.append(series2)

    # установить заголовок графика
    chart.title = "Графики исходной (синий), расчитанной выбр. (краcный), расчитанной fmin (зеленый) интенсивности"

    # установить заголовок оси X
    chart.x_axis.title = "Расстояние от гипоцентра, км"
    # установить заголовок оси Y
    cp = CharacterProperties(sz=2000)
    chart.y_axis.title = "Интенсивность"

    chart.style = 10;     chart.height = 20;    chart.width = 30

    # cp = CharacterProperties(sz=1200)
    chart_sheet.add_chart(chart, "G2")

    try:
        my_wb.save(fname)
        bres = True
    except Exception:
        bres = False
    return bres



def prepare_and_output_xlsres(resdir:str, inffname:str, imod:np.ndarray, ini_dat: list, res_dat: list,
                              x_len: np.ndarray, y_i_fact: np.ndarray, y_i_mod: np.ndarray, name_sort: np.ndarray) -> bool:
    (lat, lon, alt, ifact, di, num, the_name) = pinp_struct.get_all_arrays()
    fname = prepare_output_xlsres_fn(resdir, inffname)
    # print(fname)
    bres = output_xls_res(fname, lat , lon, alt, ifact, di, num, the_name, imod, ini_dat, res_dat, x_len, y_i_fact, y_i_mod, name_sort)
    return bres

def output_xls_res(fname:str, lat: np.ndarray, lon: np.ndarray, alt: np.ndarray, ifact: np.ndarray, di: np.ndarray, N: np.ndarray,
                the_name: np.ndarray, imod: np.ndarray, ini_dat: list, res_dat: list, x_len: np.ndarray,
                y_i_fact: np.ndarray, y_i_mod: np.ndarray, name_sort: np.ndarray) -> bool:
    """
    Вывод в xlsx файл исходных данных и результаитов минимизации:
    lat, lon, alt, ifact, di, N, the_name, imod
    """
    ################### Первый лист
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    my_sheet.title = "Данные и Imod"
    llen = len(lat)
    my_sheet.cell(row=1, column= 1).value = 'Lat'
    my_sheet.cell(row=1, column= 2).value = 'Lon'
    my_sheet.cell(row=1, column= 3).value = 'Alt'
    my_sheet.cell(row=1, column= 4).value = 'Ifact'
    my_sheet.cell(row=1, column= 5).value = 'dI'
    my_sheet.cell(row=1, column= 6).value = 'N'
    my_sheet.cell(row=1, column= 7).value = 'Нас.пункт'
    my_sheet.cell(row=1, column= 8).value = 'Imod'
    my_sheet.cell(row=1, column= 9).value = 'В пределах dI'
    my_sheet.cell(row=1, column=10).value = 'От эпицентра (выбр.точка), км'
    dist_on_surf = ptest_alg.calc_dist_onsurf_full(res_dat[1], res_dat[2], lat, lon)
    for i in range(llen):
        rrow = i + 2
        my_sheet.cell(row=rrow, column=1).value = lat[i]
        my_sheet.cell(row=rrow, column=2).value = lon[i]
        my_sheet.cell(row=rrow, column=3).value = alt[i]
        my_sheet.cell(row=rrow, column=4).value = ifact[i]
        my_sheet.cell(row=rrow, column=5).value = di[i]
        my_sheet.cell(row=rrow, column=6).value = N[i]
        my_sheet.cell(row=rrow, column=7).value = the_name[i]
        my_sheet.cell(row=rrow, column=8).value = round(imod[i], 4)
        my_sheet.cell(row=rrow, column=9).value = pfunct.get_xls_bool(abs(ifact[i]-imod[i]) <= (di[i]/2))
        my_sheet.cell(row=rrow, column=10).value = round(dist_on_surf[i], 4)

    my_sheet.column_dimensions['G'].width = 25
    my_sheet.column_dimensions['I'].width = 15
    my_sheet.column_dimensions['J'].width = 30

    ################### Второй лист
    # https://ru.stackoverflow.com/questions/828957/С-помощью-python-cоздать-новый-лист-в-уже-имеющемся-excel-файле
    res_sheet = my_wb.create_sheet('Результаты')
    res_sheet.cell(row=1, column=1).value = 'Начальное приближение'
    res_sheet.cell(row=2, column=1).value = 'Широта';     res_sheet.cell(row=2, column=2).value = ini_dat[0]

    res_sheet.cell(row=3, column=1).value = 'Долгота';    res_sheet.cell(row=3, column=2).value = ini_dat[1]

    res_sheet.cell(row=4, column=1).value = 'Глубина';    res_sheet.cell(row=4, column=2).value = ini_dat[2]

    res_sheet.cell(row=5, column=1).value = 'Магнитуда';  res_sheet.cell(row=5, column=2).value = ini_dat[3]

    res_sheet.cell(row=6, column=1).value = ' '
    res_sheet.cell(row=7, column=1).value = 'Выбранный результат минимизации'
    res_sheet.cell(row=8, column=1).value = 'Итерация (строка в файле ***_pointlist)'
    res_sheet.cell(row=8, column=2).value = res_dat[0]

    res_sheet.cell(row=9, column=1).value = 'Широта';     res_sheet.cell(row=9, column=2).value = res_dat[1]

    res_sheet.cell(row=10, column=1).value = 'Долгота';   res_sheet.cell(row=10, column=2).value = res_dat[2]

    res_sheet.cell(row=11, column=1).value = 'Глубина';   res_sheet.cell(row=11, column=2).value = res_dat[3]
    res_sheet.cell(row=12, column=1).value = 'Магнитуда'; res_sheet.cell(row=12, column=2).value = res_dat[4]
    res_sheet.cell(row=13, column=1).value = 'Значение целевой функции'
    res_sheet.cell(row=13, column=2).value = res_dat[5]
    # http://espressocode.top/python-adjusting-rows-and-columns-of-an-excel-file-using-openpyxl-module/
    res_sheet.column_dimensions['A'].width = 40
    ################### Третий лист
    chart_sheet = my_wb.create_sheet('Диаграмма')
    chart_sheet.cell(row=1, column=1).value = 'От гипоцентра, км'
    chart_sheet.cell(row=1, column=2).value = 'I_fact'
    chart_sheet.cell(row=1, column=3).value = 'I_mod'
    chart_sheet.cell(row=1, column=4).value = 'Нас.пункт'
    nnn = len(x_len)
    for i in range(nnn):
        rrow = i + 2
        chart_sheet.cell(row=rrow, column=1).value = x_len[i]
        chart_sheet.cell(row=rrow, column=2).value = y_i_fact[i]
        chart_sheet.cell(row=rrow, column=3).value = y_i_mod[i]
        chart_sheet.cell(row=rrow, column=4).value = name_sort[i]

    chart_sheet.column_dimensions['A'].width = 20
    chart_sheet.column_dimensions['B'].width = 10
    chart_sheet.column_dimensions['C'].width = 10
    chart_sheet.column_dimensions['D'].width = 20

    chart = ScatterChart()
    # создать данные для построения графика
    xvalues = Reference(chart_sheet, min_col=1, min_row=2, max_row=nnn+2)
    yvalues = Reference(chart_sheet, min_col=2, min_row=2, max_row=nnn+2)
    y1values = Reference(chart_sheet, min_col=3, min_row=2, max_row=nnn+2)
    size = Reference(chart_sheet, min_col=3, min_row=2, max_row=nnn+2)

    # создать серии данных
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="Исходная")
    series.graphicalProperties.line.solidFill = '0000FF'
    series.graphicalProperties.line.width = 25000
    series1 = Series(values=y1values, xvalues=xvalues, zvalues=size, title="Рассчитанная")
    series1.graphicalProperties.line.solidFill = 'FF0000'
    series1.graphicalProperties.line.width = 25000

    # https://stackoverflow.com/questions/56551838/with-python-openpyxl-how-do-you-change-the-font-size-in-a-chart-legend
    font_test = openpyxl.drawing.text.Font(typeface='Arial')
    cp = CharacterProperties(latin=font_test, sz=1500)
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    chart.x_axis.scaling.min = round(math.floor(np.min(x_len)),-1)-10
    chart.x_axis.scaling.max = round(math.ceil(np.max(x_len)),-1)+10
    # chart.y_axis.title.

    chart.y_axis.scaling.min = round(math.floor(min(np.min(y_i_fact), np.min(y_i_mod))),0)-1
    dat = round(math.ceil(max(np.max(y_i_fact), np.max(y_i_mod))),0)+1
    chart.y_axis.scaling.max = dat
    chart.legend.position = 'b'

    # добавить данные серии в объект диаграммы
    chart.series.append(series)
    chart.series.append(series1)

    # установить заголовок графика
    chart.title = "Графики исходной (синий) и расчитанной (краcный) интенсивности"

    # установить заголовок оси X
    chart.x_axis.title = "Расстояние от гипоцентра, км"
    # установить заголовок оси Y
    cp = CharacterProperties(sz=2000)
    chart.y_axis.title = "Интенсивность"

    chart.style = 10;     chart.height = 20;    chart.width = 30

    cp = CharacterProperties(sz=1200)

    # добавить диаграмму на лист верхний левый угол графика привязан к ячейке E2.
    chart_sheet.add_chart(chart, "F2")

    try:
        my_wb.save(fname)
        bres = True
    except Exception:
        bres = False
    return bres


def output_xls_pointlist(fname: str, result_list: list) -> None:
    """
    Вывод в xlsx файл итераций минимизации:
    номер, lat, lon, dep, mag, fun, dLat %, dLon %, dDep %, dMag %, dFun %
    """
    # Чтение и запись в файл Excel с использованием модуля Python openpyxl
    # https://pythononline.ru/question/chtenie-i-zapis-v-fayl-excel-s-ispolzovaniem-modulya-python-openpyxl
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    my_sheet.title = "Результаты"
    llen = len(result_list)
    my_sheet.cell(row=1, column= 1).value = 'i'
    my_sheet.cell(row=1, column= 2).value = 'Lat'
    my_sheet.cell(row=1, column= 3).value = 'Lon'
    my_sheet.cell(row=1, column= 4).value = 'Dep'
    my_sheet.cell(row=1, column= 5).value = 'Mag'
    my_sheet.cell(row=1, column= 6).value = 'Fun'
    my_sheet.cell(row=1, column= 7).value = 'dLat, %'
    my_sheet.cell(row=1, column= 8).value = 'dLon, %'
    my_sheet.cell(row=1, column= 9).value = 'dDep, %'
    my_sheet.cell(row=1, column=10).value = 'dMag, %'
    my_sheet.cell(row=1, column=11).value = 'dFun, %'
    my_sheet.cell(row=1, column=12).value = 'max_var,%/dFun,%'
    for i in range(llen):
        d = result_list[i]
        rrow = i + 2
        my_sheet.cell(row=rrow, column=1).value = i  # i
        my_sheet.cell(row=rrow, column=2).value = d[0]  # Lat
        my_sheet.cell(row=rrow, column=3).value = d[1]  # Lon
        my_sheet.cell(row=rrow, column=4).value = d[2]  # Dep
        my_sheet.cell(row=rrow, column=5).value = d[3]  # Mag
        my_sheet.cell(row=rrow, column=6).value = d[4]  # Fun
        if rrow > 2:
            my_sheet.cell(row=rrow, column= 7).value = d[5]  # dLat
            my_sheet.cell(row=rrow, column= 8).value = d[6]  # dLon
            my_sheet.cell(row=rrow, column= 9).value = d[7]  # dDep
            my_sheet.cell(row=rrow, column=10).value = d[8]  # dMag
            my_sheet.cell(row=rrow, column=11).value = d[9]  # dFun
            my_sheet.cell(row=rrow, column=12).value = d[10]  # max(dvar)/dFun

    my_wb.save(fname)

                                         # num, lat_,  lon_,  dep_,  mag_,  fun_
def get_true_result(result_list: list) -> (int, float, float, float, float, float):
    """
    Выборка из result_list
    """
    global gran
    llen = len(result_list)
    # print('llen =', llen)
    num = 0
    for i in range(1, llen):
        d = result_list[i]
        f_ = d[4]
        df_ = abs(d[9])
        if ((i != 1) and (df_ < gran)) or (f_ < gran):
            num = i
            break

    d = result_list[num]
    lat_ = d[0]  # Lat
    lon_ = d[1]  # Lon
    dep_ = d[2]  # Dep
    mag_ = d[3]  # Mag
    fun_ = d[4]  # Fun

    return num, lat_,  lon_,  dep_,  mag_,  fun_


# @numba.njit
def minimize_func(n: int, lat_arr: np.ndarray, lon_arr: np.ndarray,
                           h_Arr: np.ndarray, i_fact_arr: np.ndarray, di_arr: np.ndarray) -> (int, float, float, float, float, float, list):
    global log_file_name
    a: float; b: float; c: float
    str_res_: str
    f1: float
    proc_f: float  # процент изменения f на каждой итерации
    iii: int = 0
    result_list: list= []

    def main_objective_function(x0) -> float:
        nonlocal iii
        nonlocal result_list
        f1 = pinp_struct.objective_function(n, lat_arr, lon_arr, h_Arr, i_fact_arr, di_arr, x0[0], x0[1], x0[2], x0[3], a, b, c)

        # вычисление доп.параметров для останова приращений переменных и функции и dgran_ - спец.итоговой функции
        (dx0, dx1, dx2, dx3, df, dgran_) = calc_diff(iii, x0[0], x0[1], x0[2], x0[3], f1, result_list)
        result_list.append([x0[0], x0[1], x0[2], x0[3], f1,
                              dx0, dx1,   dx2,   dx3,  df, dgran_])
        # print('%5d lat_ %10.6f =   lon_ =  %10.6f   dep_ = %10.6f   mag_ = %10.6f   f = %10.6f' % (iii, x0[0], x0[1], x0[2], x0[3], f1))
        # print('      dlat_ = %10.6f  dlon_ =  %10.6f  ddep_ = %10.6f  dmag_ = %10.6f  df = %10.6f' % (dx0, dx1, dx2, dx3, df))
        # print('           %10.6f            %10.6f           %10.6f          %10.6f       %10.6f' % (dx0, dx1, dx2, dx3, df))
        iii += 1
        return f1

    # scipy.optimize.minimize
    # If callback returns True the algorithm execution is terminated
    # https://docs.scipy.org/doc/scipy/reference/generated/scipy.optimize.minimize.html
    # Python3.4 scipy.optimize.minimize callbacks multiple times the objective function in one iteration
    # https://stackoverflow.com/questions/32477371/python3-4-scipy-optimize-minimize-callbacks-multiple-times-the-objective-functio
    def call_back_f(Xi) -> bool:
        nonlocal iii
        nonlocal result_list

        res_analyz: bool = analyze_diff(iii-1, result_list)
        return res_analyz

    (a, b, c) = pinp_struct.get_a_b_c(pinp_struct.curr_nstruct)
    (ini_lat_, ini_lon_, ini_dep_, ini_mag_) = pinp_struct.get_ini(pinp_struct.curr_nstruct)
    #               x0[0]      x0[1]     x0[2]     x0[3]
    x0 = np.array([ini_lat_, ini_lon_, ini_dep_, ini_mag_])
    # print('x0 = ', x0);   sys.exit()
    f = main_objective_function
    # res = minimize(f, x0, method='nelder-mead', tol=1e-10,
    #                options={'maxiter': 80000, 'xatol': 0.00001, 'fatol': 1e-12, 'adaptive': True})
    # res = minimize(f, x0, method='nelder-mead', tol=1e-10)

    res = minimize(f, x0, method='nelder-mead', callback=call_back_f)
    # res = minimize(f, x0, method='nelder-mead', options={'maxiter': 8000,'fatol': 0.01, 'adaptive': True})

    # вывод всей последовательности точек минимизации в файл
    # [если истина] if [выражение] else [если ложь]
    if pinp_struct.is_txt_res_file:
        output_txt_res(log_file_name, result_list)
    else:
        output_xls_pointlist(log_file_name, result_list)

    # print('Начальные значения')
    # print('ini_lat_ = %7.3f' % ini_lat_)
    # print('ini_lon_ = %7.3f' % ini_lon_)

    # print('Результат автоматический')
    # print('res.x', res.x)
    # print('lat = res.x[0] = %8.5f' % res.x[0])
    # print('lon = res.x[1] = %8.5f' % res.x[1])
    # print('dep = res.x[2] = %8.5f' % res.x[2])
    # print('mag = res.x[3] = %8.5f' % res.x[3])
    # print('res.success = ', res.success)  # bool - Whether or not the optimizer exited successfully.
    # print('res.message = ', res.message)  # str - Description of the cause of the termination.
    # print('res.fun = ', res.fun)  # fun - Values of objective function
    # print('res.nfev = ', res.nfev, ' main_objective_function, число вычислений')      # Number of evaluations of the objective functions
    # print('res.nit = ', res.nit)      # int - Number of iterations performed by the optimizer
    # print(f.__name__+' = ', f(res.x))  # main_objective_function

    (num, lat_, lon_, dep_, mag_, fun_) = get_true_result(result_list)

    return num, lat_, lon_, dep_, mag_, fun_, result_list

def create_str_res(name_sq, ini_lat_, ini_lon_, ini_dep_, ini_mag_,
                   num, lat_, lon_, dep_, mag_, fun_) -> str:
    as_ = ' '*5

    str_res_ = '\n'
    str_res_ += (as_+name_sq).center(45)
    str_res_ += '\n'
    str_res_ += '\n'
    str_res_ += as_+'        Начальное приближение\n'
    str_res_ += as_+'     Широта = '+format(ini_lat_,'7.3f') + '\n'
    str_res_ += as_+'    Долгота = '+format(ini_lon_,'7.3f') + '\n'
    str_res_ += as_+'    Глубина = '+format(ini_dep_,'7.3f') + '\n'
    str_res_ += as_+'  Магнитуда = '+format(ini_mag_,'7.3f') + '\n'+ '\n'

    str_res_ += as_+'        Выбранный результат минимизации' + '\n'
    str_res_ += as_+'   Итерация (строка в файле результата) = ' + format(num,'>4d') + '\n'
    str_res_ += as_+'       Широта = '+format(lat_,'7.3f') + '\n'
    str_res_ += as_+'      Долгота = '+format(lon_,'7.3f') + '\n'
    str_res_ += as_+'      Глубина = '+format(dep_,'7.3f') + '\n'
    str_res_ += as_+'    Магнитуда = '+format(mag_,'7.3f') + '\n'
    str_res_ += as_+' Значение целевой функции = '+format(fun_,'7.3f') + '\n'
    return str_res_


def create_str_res_calc2(name_sq, lat_, lon_, ini_dep_, ini_mag_,  dep_, mag_, corr_dep, corr_mag,
                         min_sum, min_sum_correct, info_str) -> str:
    as_ = ' '*5

    str_res_ = '\n'
    str_res_ += (as_+name_sq)
    str_res_ += '\n'
    str_res_ += '\n'
    str_res_ += as_+'     Широта = '+format(lat_,'7.3f') + '\n'
    str_res_ += as_+'    Долгота = '+format(lon_,'7.3f') + '\n' + '\n'
    str_res_ += '\n'
    str_res_ += as_+'        Начальное приближение\n'
    str_res_ += as_+'      Глубина = '+format(ini_dep_,'7.3f') + '\n'
    str_res_ += as_+'    Магнитуда = '+format(ini_mag_,'7.3f') + '\n'
    str_res_ += '\n'
    str_res_ += as_+'        Наименьшее значение целевой функции' + '\n'
#   str_res_ += as_+'   Номер решения (строка в файле результата) = ' + format(min_i,'>5d') + '\n'
    str_res_ += as_+'      Глубина = '+format(dep_,'7.3f') + '\n'
    str_res_ += as_+'    Магнитуда = '+format(mag_,'7.3f') + '\n'
    str_res_ += as_+' Значение целевой функции = '+format(min_sum,'10.5f') + '\n'
    str_res_ += '\n'
    str_res_ += as_+'        Выбраный результат минимизации' + '\n'
#   str_res_ += as_+'   Номер решения (строка в файле результата) = ' + format(min_icorrect,'>5d') + '\n'
    str_res_ += as_+'      Глубина = '+format(corr_dep,'7.3f') + '\n'
    str_res_ += as_+'    Магнитуда = '+format(corr_mag,'7.3f') + '\n'
    str_res_ += as_+' Значение целевой функции = '+format(min_sum_correct,'10.5f') + '\n'
    str_res_ += as_+info_str+ '\n'
    return str_res_


def add_info():
    lat_ = 55.92745139505674;  lon_ = 57.33186006061451; dep_ = 29.999999999999993; mag_ = 5.723468511118773
    print('N результата 1731')
    # N1731 res.fun =  44.16925687182368
    result_control(lat_, lon_, dep_, mag_)

    print('N  результата 35')
    # N35 res.fun = 46.196595366820034
    lat_ = 55.860328095037005; lon_  = 57.002550801519476;  dep_ = 10.062189981696722;  mag = 5.612690867989407
    result_control(lat_, lon_, dep_, mag_)

    print('N  результата 51')
    # N51 res.fun = 44.99874411859368
    lat_ = 55.84245014011909;  lon_ = 57.22878737017246;  dep_ = 10.172387821353635; mag_ = 5.699159301892644
    result_control(lat_, lon_, dep_, mag_)
